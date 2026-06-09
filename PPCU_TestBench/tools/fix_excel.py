"""Apply structure improvements to PPCU Telemetry Excel."""
import os, tempfile, copy
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

SRC = "E:/Codex Workspace/01_Projects/项目文档/通讯协议/excel版遥测大表/PPCU_Telemetry_Database_v2.xlsx"

def parse_enum_to_decimal(s):
    """Convert enum string to decimal format: '00b:关; 11b:开' -> '0:关; 3:开'"""
    if not s: return ""
    parts = str(s).replace("，",";").replace("；",";").replace("\n",";").strip().split(";")
    result = []
    for part in parts:
        part = part.strip()
        if not part: continue
        if ":" in part:
            key, val = part.split(":", 1)
            key = key.strip()
            val = val.strip()
            if not key and not val: continue
            if not key: result.append(val); continue
            try:
                if key.endswith("b"): k = int(key[:-1], 2)
                elif key.endswith("H"): k = int(key[:-1], 16)
                elif key == "其他" or key == "其它": continue
                else: k = int(key, 0)
                result.append(f"{k}:{val}")
            except ValueError:
                result.append(f"{key}:{val}")
        else:
            part = part.strip()
            if part: result.append(part)
    return "; ".join(result)

def fix_excel():
    wb = openpyxl.load_workbook(SRC)
    
    for sname in ["TM1 参数表", "TM2 参数表", "查询包 参数表"]:
        ws = wb[sname]
        print(f"\nProcessing {sname}...")
        
        # Insert new column for data_offset after column D (bit_offset)
        ws.insert_cols(5)  # Insert at column E
        ws.cell(1, 5, value="数据域偏移\ndata_offset")
        ws.cell(1, 5).font = Font(bold=True, color="FFFFFF")
        ws.cell(1, 5).fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        ws.cell(1, 5).alignment = Alignment(horizontal="center", wrap_text=True)
        
        changes = {"enum_fixed": 0, "dec_filled": 0, "type_fixed": 0, "range_filled": 0}
        for r in range(2, ws.max_row + 1):
            pid = ws.cell(r, 2).value
            if not pid: continue
            
            # Calculate and fill data_offset
            bo_raw = ws.cell(r, 4).value  # bit_offset (col D)
            if bo_raw is not None:
                try: bo_val = int(str(bo_raw).strip(), 0)
                except: bo_val = 0
                do = bo_val // 8 - 8
                ws.cell(r, 5).value = do  # data_offset at col E
            
            bl = ws.cell(r, 6).value  # bit_length (was col 5, now col 6)
            ptype = str(ws.cell(r, 7).value or "").strip()  # type (was col 6, now col 7)
            scale = ws.cell(r, 9).value  # scale (was col 8, now col 9)
            dec = ws.cell(r, 10).value   # decimal (was col 9, now col 10)
            rmin = ws.cell(r, 12).value  # range_min (was col 11, now col 12)
            rmax = ws.cell(r, 13).value  # range_max (was col 12, now col 13)
            enum_str = ws.cell(r, 14).value  # enum desc (was col 13, now col 14)
            
            # Fix type: change ambiguous 'float' for 16-bit to 'uint16'
            if ptype.lower() == "float":
                if bl == 16:
                    ws.cell(r, 7).value = "uint16"
                    changes["type_fixed"] += 1
                elif bl == 32:
                    ws.cell(r, 7).value = "float32"
                    changes["type_fixed"] += 1
            
            # Fill missing decimal_places for scaled values
            if scale and isinstance(scale, (int, float)) and 0 < float(scale) < 1:
                if dec is None or str(dec).strip() == "":
                    ws.cell(r, 10).value = 2
                    changes["dec_filled"] += 1
            
            # Standardize enum format
            if "enum" in ptype.lower() and enum_str:
                old = str(enum_str).strip()
                new = parse_enum_to_decimal(old)
                if new != old:
                    ws.cell(r, 14).value = new
                    changes["enum_fixed"] += 1
            
            # Add range for voltage/current/sensor types if missing
            name = str(ws.cell(r, 3).value or "")
            if rmin is None or str(rmin).strip() == "":
                if "电压" in name and "300" in str(rmax or ""):
                    ws.cell(r, 12).value = 0
                    changes["range_filled"] += 1
                elif "电流" in name and rmax is None:
                    ws.cell(r, 13).value = 30
                    changes["range_filled"] += 1
        
        print(f"  Enums fixed: {changes['enum_fixed']}")
        print(f"  Decimal places filled: {changes['dec_filled']}")
        print(f"  Types fixed: {changes['type_fixed']}")
        print(f"  Ranges filled: {changes['range_filled']}")
    
    # Save
    out = os.path.join(tempfile.gettempdir(), "PPCU_Telemetry_Database_v3.xlsx")
    wb.save(out)
    print(f"\nSaved to: {out}")
    print(f"Copy with: Copy-Item $env:TEMP\\PPCU_Telemetry_Database_v3.xlsx ...")

if __name__ == "__main__":
    fix_excel()
