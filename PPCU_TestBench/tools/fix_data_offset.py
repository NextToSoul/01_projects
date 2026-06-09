"""Fix data_offset column values: read bit_offset from col 5, write to col 4."""
import os, tempfile, shutil
import openpyxl
from pathlib import Path

SRC = Path("E:/Codex Workspace/01_Projects/项目文档/通讯协议/excel版遥测大表/PPCU_Telemetry_Database_v3.xlsx")
tmp = os.path.join(tempfile.gettempdir(), "PPCU_v3_fix.xlsx")
shutil.copy2(str(SRC), tmp)
wb = openpyxl.load_workbook(tmp, data_only=True)

for sname in ["TM1 参数表", "TM2 参数表", "查询包 参数表"]:
    ws = wb[sname]
    count = 0
    for r in range(2, ws.max_row + 1):
        pid = ws.cell(r, 2).value
        if not pid: continue
        bo_raw = ws.cell(r, 5).value  # bit_offset at col 5
        if bo_raw is not None and isinstance(bo_raw, (int, float)):
            bo_val = int(str(bo_raw).strip(), 0)
            do = bo_val // 8 - 8
            ws.cell(r, 4).value = do  # data_offset at col 4
            count += 1
    print(f"{sname}: {count} data_offset values fixed")

out = os.path.join(tempfile.gettempdir(), "PPCU_v3_fixed.xlsx")
wb.save(out)
print(f"\nSaved to: {out}")
print(f"Copy: Copy-Item $env:TEMP\\PPCU_v3_fixed.xlsx ...")
