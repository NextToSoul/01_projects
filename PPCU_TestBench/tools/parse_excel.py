"""Parse PPCU Telemetry Excel and generate YAML definitions."""
import openpyxl, json
from pathlib import Path

SRC = Path(__file__).resolve().parent.parent.parent / "项目文档" / "通讯协议" / "excel版遥测大表" / "PPCU_Telemetry_Database_v2.xlsx"

def dump():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"\n=== {name} === ({ws.max_row} rows x {ws.max_column} cols)")
        for r in range(1, min(ws.max_row + 1, 60)):
            vals = []
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                vals.append(str(v).replace(chr(10), " ") if v is not None else "")
            if any(vals):
                print(f"  R{r}: " + " | ".join(vals))

if __name__ == "__main__":
    dump()
"""Parse PPCU Telemetry Excel and generate YAML + verify with real frames."""
import openpyxl, struct, sys
from pathlib import Path

SRC = Path(__file__).resolve().parent.parent.parent
EXCEL = SRC / "项目文档" / "通讯协议" / "excel版遥测大表" / "PPCU_Telemetry_Database_v2.xlsx"

# Real response frames from user
TM1_FRAME = bytes.fromhex("1ACF0525C003002303000000000F0001CF000000E10270026001F5017D01FA0B4F0000088600000000000001F900")
TM2_FRAME = bytes.fromhex("1ACF0526C0010043650000000000000000000001040FA5000000000000000000000700000000140002000000000000579F810CA38360FC000000780033006400040708659A000000F76F")
QRY_FRAME = bytes.fromhex("1ACF0521C001018D4370000042F00000439600000000012C4370000041200000453B800044E1000044E1000045A8C0004416000043960000441600004270000040A0000042700000424800000000001E00000002427000003F19999A3D0106253BA3D70A3B03126F404000003F8000003E4CCCCD3CA3D70A3E4CCCCD00000002BC47E282BCAC258DBCC97311BC36A61A3CD4C76D3DABEF073E1876193E4D0F1F3E6113C7BC47E282BCAC258DBCC97311BC36A61A3CD4C76D3DABEF073E1876193E4D0F1F3E6113C7BC47E282BCAC258DBCC97311BC36A61A3CD4C76D3DABEF073E1876193E4D0F1F3E6113C73E873A4E3CD955A23F974CE24214B4753F602BFA3E0C06E23F8000003F8000003F8000003F19999A3F3333333F8000003E99999A3ECCCCCD440C000042F000003FA666663ECCCCCD3F4CCCCD3FC0000040000000402000004039999A404CCCCD40600000438C0000437A00003FD9999A3DF5C28F4220000040A000004019999A3E99999A3E99999A3F8000000064000A000003E805DC03E800500014000001F400640050000000000000813A")

def compute_cs(frame):
    total = sum(frame[2:-2]) & 0xFFFF
    return ((~total) & 0xFFFF).to_bytes(2, "big")

def parse_frame(frame, name):
    hdr = frame[:2]
    apid = frame[2:4]
    seq = frame[4:6]
    data_len = int.from_bytes(frame[6:8], "big")
    computed = compute_cs(frame)
    actual = frame[-2:]
    body = frame[8:8+data_len-2]
    print(f"\n--- {name} ---")
    print(f"  Header: {hdr.hex()} APID: {apid.hex()} Seq: {seq.hex()}")
    print(f"  Length field: {data_len} ({data_len:#06x})")
    print(f"  Total frame: {len(frame)}B, Data field: {data_len}B")
    print(f"  Body (telemetry): {len(body)}B")
    print(f"  Checksum: actual={actual.hex()} computed={computed.hex()} MATCH={actual==computed}")
    return body, data_len

if __name__ == "__main__":
    b1, l1 = parse_frame(TM1_FRAME, "TM1")
    b2, l2 = parse_frame(TM2_FRAME, "TM2")
    b3, l3 = parse_frame(QRY_FRAME, "Query")
    
    print(f"\nTM1 body hex: {b1.hex().upper()}")
    print(f"TM2 body hex: {b2.hex().upper()}")
    print(f"QRY body hex: {b3.hex().upper()}")
"""Parse PPCU Telemetry Excel and generate YAML."""
import openpyxl
from pathlib import Path

EXCEL = (Path(__file__).resolve().parent.parent.parent /
    "项目文档" / "通讯协议" / "excel版遥测大表" / "PPCU_Telemetry_Database_v2.xlsx")

def bit_to_do(bit_off):
    """bit_offset (frame-level) -> data_offset, bit_in_byte"""
    byte_pos = bit_off // 8
    return byte_pos - 8, bit_off % 8  # data_offset, bit_in_byte

def fmt_yaml(params):
    lines = []
    for p in params:
        lines.append(f"  - id: {p['pid']}")
        lines.append(f"    name: {p['name']}")
        lines.append(f"    byte_offset: {p['byte_offset']}")
        if "bit_offset" in p:
            lines.append(f"    bit_offset: {p['bit_offset']}")
            lines.append(f"    bit_length: {p['bit_length']}")
        else:
            lines.append(f"    byte_length: {p['byte_length']}")
        lines.append(f"    type: {p['type']}")
        if p.get("endian"):
            lines.append(f"    endian: {p['endian']}")
        lines.append(f"    scale: {p['scale']}")
        lines.append(f"    decimal_places: {p['decimal_places']}")
        if p.get("unit"):
            lines.append(f"    unit: {p['unit']}")
        if "range_min" in p:
            lines.append(f"    range_min: {p['range_min']}")
            lines.append(f"    range_max: {p['range_max']}")
        lines.append("")
    return "\n".join(lines)

def gen(sheet_name, tm_prefix):
    ws = openpyxl.load_workbook(EXCEL, data_only=True)[sheet_name]
    params = []
    for r in range(2, ws.max_row + 1):
        pid = ws.cell(r, 2).value
        name = ws.cell(r, 3).value
        bo = ws.cell(r, 4).value  # bit_offset
        bl = ws.cell(r, 5).value  # bit_length
        ptype = ws.cell(r, 6).value
        scale = ws.cell(r, 8).value or 1.0
        dec = ws.cell(r, 9).value
        unit = ws.cell(r, 10).value or ""
        rmin = ws.cell(r, 11).value
        rmax = ws.cell(r, 12).value
        if pid is None or name is None or bo is None:
            continue
        do, bib = bit_to_do(bo)  # data_offset, bit_in_byte
        if do < 0:
            continue
        p = {"pid": pid.strip(), "name": str(name).strip(),
             "byte_offset": do, "scale": float(scale),
             "decimal_places": int(dec) if dec else 0}
        if bl in (1,2,3,4,5,6,7):
            p["bit_offset"] = bib
            p["bit_length"] = bl
            typ = str(ptype or "").lower()
            p["type"] = "enum" if "enum" in typ else "uint8"
        elif bl == 8:
            p["byte_length"] = 1
            p["endian"] = "big"
            typ = str(ptype or "").lower()
            p["type"] = "enum" if "enum" in typ else "uint8"
        elif bl == 16:
            p["byte_length"] = 2
            p["endian"] = "big"
            typ = str(ptype or "").lower()
            p["type"] = "enum" if "enum" in typ else ("uint16" if "float" not in typ else "uint16")
        elif bl == 24:
            p["byte_length"] = 3
            p["endian"] = "big"
            p["type"] = "uint32"
        elif bl == 32:
            p["byte_length"] = 4
            p["endian"] = "big"
            typ = str(ptype or "").lower()
            p["type"] = "float32" if "float" in typ else "uint32"
        if rmin is not None: p["range_min"] = int(rmin)
        if rmax is not None: p["range_max"] = int(rmax)
        if unit: p["unit"] = str(unit).strip()
        params.append(p)
    return params

if __name__ == "__main__":
    for sname, prefix in [("TM1 参数表","tm1"),("TM2 参数表","tm2"),("查询包 参数表","query")]:
        params = gen(sname, prefix)
        yml = fmt_yaml(params)
        print(f"=== {sname}: {len(params)} parameters ===\n")
        header = f"tm_type: {prefix}\ndescription: {sname.replace(chr(39),'')}\ncycle_s: 1\n\nparams:\n"
        print(header + yml)
