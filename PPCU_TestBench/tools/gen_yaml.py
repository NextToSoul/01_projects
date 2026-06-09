# -*- coding: utf-8 -*-
"""Generate telemetry YAML from PPCU_Telemetry_Database_v4.xlsx."""
import os, sys, tempfile, shutil
import openpyxl
from pathlib import Path

SRC = Path("E:/Codex Workspace/01_Projects/项目文档/通讯协议/excel版遥测大表/PPCU_Telemetry_Database_v4.xlsx")
TMP = os.path.join(tempfile.gettempdir(), "ppcu_v4_gen.xlsx")
OUT = tempfile.gettempdir()

try:
    shutil.copy2(str(SRC), TMP)
except:
    TMP = "C:/Users/19757/AppData/Local/Temp/PPCU_v4.xlsx"

WS_NAMES = {"tm1": "TM1 参数表", "tm2": "TM2 参数表", "query": "查询包 参数表"}

def safe_int(v):
    if v is None: return None
    if isinstance(v, (int, float)): return int(v)
    try: return int(str(v), 0)
    except: return None

def safe_float(v):
    if v is None: return 1.0
    try: return float(v)
    except: return 1.0

def gen(prefix):
    sname = WS_NAMES[prefix]
    ws = openpyxl.load_workbook(TMP, data_only=True)[sname]
    # v4 columns: A=序号, B=ID, C=名称, D=data_offset, E=bit_offset,
    # F=bit_length, G=type, H=endian, I=scale, J=decimal, K=unit,
    # L=range_min, M=range_max, N=enum_desc
    lines = ["tm_type: " + prefix, "description: " + sname.replace(chr(39),""), "params:"]
    cnt = 0
    for r in range(2, ws.max_row + 1):
        pid = ws.cell(r, 2).value
        if not pid: continue
        do_raw = ws.cell(r, 4).value  # data_offset col D
        do = safe_int(do_raw)
        if do is None: continue
        cnt += 1; lines.append("")
        lines.append("  - id: " + str(pid).strip())
        lines.append("    name: " + str(ws.cell(r, 3).value or "").strip())
        lines.append("    byte_offset: " + str(do))
        bl = ws.cell(r, 6).value  # bit_length col F
        ptype = str(ws.cell(r, 7).value or "").lower()  # type col G
        bo_raw = ws.cell(r, 5).value  # bit_offset col E
        is_enum = "enum" in ptype
        if bl in (1,2,3,4,5,6,7):
            bib = int(bo_raw) % 8 if bo_raw is not None else 0
            bib = 8 - bib - bl  # MSB (Excel) -> LSB (code) conversion
            lines.append("    bit_offset: " + str(bib))
            lines.append("    bit_length: " + str(bl))
            lines.append("    type: enum" if is_enum else "    type: uint8")
        else:
            lines.append("    byte_length: " + str(bl // 8 if bl and bl > 8 else (bl if bl else 1)))
            if bl == 32 and "float" in ptype:
                lines.append("    type: float32"); lines.append("    scale: 1.0")
            else:
                t = "uint16" if bl == 16 else ("uint32" if bl in (24,32) else "uint8")
                lines.append("    type: " + t)
                lines.append("    scale: " + str(safe_float(ws.cell(r, 9).value)))
            lines.append("    endian: big")
        dec = ws.cell(r, 10).value
        lines.append("    decimal_places: " + str(int(dec) if dec else 0))
        unit = ws.cell(r, 11).value or ""
        if unit: lines.append("    unit: " + str(unit).strip())
        rmin_i = safe_int(ws.cell(r, 12).value)
        rmax_i = safe_int(ws.cell(r, 13).value)
        if rmin_i is not None: lines.append("    range_min: " + str(rmin_i))
        if rmax_i is not None: lines.append("    range_max: " + str(rmax_i))
    yml = "\n".join(lines) + "\n"
    out_name = "telemetry_" + prefix + ".yaml"
    out_path = os.path.join(OUT, out_name)
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(yml)
    print(f"Wrote {out_name}: {cnt} params")

if __name__ == "__main__":
    for p in ["tm1", "tm2", "query"]:
        try: gen(p)
        except Exception as e: print(f"Error: {p}: {e}")
